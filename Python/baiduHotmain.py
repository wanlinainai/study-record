import requests
from bs4 import BeautifulSoup
import openpyxl
import matplotlib.pyplot as plt

# requests库用于发送HTTP请求获取网页内容
# BeautifulSoup库用于解析HTML页面的内容
# openpyxl用于创建和操作Excel文件

def main():
    url = "https://top.baidu.com/board?tab=realtime"
    response = requests.get(url)
    html = response.content
    # 使用BeautifulSoup进行解析HTML
    soup = BeautifulSoup(html, "html.parser")

    # 热搜数据
    hot_searches = []
    for item in soup.find_all('div', {'class': 'c-single-text-ellipsis'}):
        hot_searches.append(item.text)

    # 保存热搜数据到Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Baidu Hot Searches'

    # 设置标题
    sheet.cell(row=1, column=1, value='百度热搜排行榜')

    # 写入热搜数据
    for i in range(len(hot_searches)):
        sheet.cell(row=i+2, column=1, value=hot_searches[i])

    workbook.save('百度热搜.xlsx')
    print('保存完成')
    # 进行可视化操作
    visualization()

## 图表可视化
def visualization():
    url = "https://top.baidu.com/board?tab=realtime"
    response = requests.get(url)
    html = response.content

    soup = BeautifulSoup(html, "html.parser")

    hot_searches = []
    for item in soup.find_all('div', {'class': 'c-single-text-ellipsis'}):
        hot_searches.append(item.text)

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    # 绘制条形图
    plt.figure(figsize=(15, 10))
    x = range(len(hot_searches))
    y = list(reversed(range(1, len(hot_searches) + 1)))
    plt.barh(x, y, tick_label=hot_searches, height=0.8) # 调整条形图的高度

    # 添加标题和标签
    plt.title('百度热搜排行榜')
    plt.xlabel('排名')
    plt.ylabel('关键词')

    # 调整坐标轴刻度
    plt.xticks(range(1, len(hot_searches) + 1))

    # 调整条形图之间的间距
    plt.subplots_adjust(hspace=0.8, wspace=0.5)

    # 显示图形
    plt.tight_layout()
    plt.show()

if __name__ == '__main__':
    main()
    print('解析完成>>>')